#!/usr/bin/env python
"""
update_kearsley_from_dfes.py
–––––––––––––––––––––––––––––
• Reads the 2024 / “Best View” column from the three DFES tables
      – Maximum Demand (MVA)  →  Sd_max   (load sheet)
      – Minimum Demand (MVA)  →  Sd_min   (load sheet)
      – Generation (MW)       →  Pg_max   (gen  sheet, summed over techs)
• Matches rows via *exactly* the text in the new “DFES Name” column.
• Writes the numbers into Kearsley_GSP_group_only.xlsx **in place**.
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# ----------------------------------------------------------------------
# Paths – change these three lines if you renamed the files/dirs
# ----------------------------------------------------------------------
DFES_WB   = Path("Input_Data\DFES_Projections\dfes_2024_main_workbook.xlsm")
TARGET_WB = Path("Input_Data\Manchester_Distribution_Network\Kearsley_GSP_group_only - Copy.xlsx")
YEAR      = 2024          # which column to pull from the DFES tables
# ----------------------------------------------------------------------

# -------------------- 1. Load DFES sheets once ------------------------
print("Loading DFES workbook …")
def read_df(sheet, hdr_row=9):
    """Return DFES sheet as DataFrame with a clean ‘Name’ index."""
    df = pd.read_excel(DFES_WB, sheet_name=sheet, header=hdr_row)
    df = df.rename(columns={"Primary": "Name"})          # harmonise header
    df = df[~df["Name"].isna()]                          # drop blank rows
    df["Name"] = df["Name"].str.strip()
    return df.set_index("Name")

df_max = read_df("Maximum Demand")
df_min = read_df("Minimum Demand")
df_gen = read_df("Generation")

# ▸ group ‘Generation’ because one PS/BSP can have several tech rows
gen_2024 = (
    df_gen.groupby(level=0)[YEAR]
          .sum()
          .astype(float)        # MW
)

# demand tables are already one line per PS; pick the YEAR column
max_2024 = df_max[YEAR].astype(float)    # MVA
min_2024 = df_min[YEAR].astype(float)    # MVA

# ------------- 2. Open the target workbook for in-place edit ----------
print("Opening Kearsley workbook …")
wb        = load_workbook(TARGET_WB)
ws_load   = wb["load"]
ws_gen    = wb["gen"]

def col_idx(ws, col_name):
    """Return 1-based column index whose header row (row 1) == col_name."""
    for j, cell in enumerate(ws[1], 1):
        if str(cell.value).strip() == col_name:
            return j
    raise ValueError(f"Column “{col_name}” not found in sheet “{ws.title}”")

# column positions (only looked up once)
LOAD_DFES_COL  = col_idx(ws_load, "DFES Name")
LOAD_Sd_max    = col_idx(ws_load, "Sd_max")
LOAD_Sd_min    = col_idx(ws_load, "Sd_min")

GEN_DFES_COL   = col_idx(ws_gen,  "DFES Name")
GEN_Pg_max     = col_idx(ws_gen,  "Pg_max")

# ---------------- 3. Populate the ‘load’ sheet ------------------------
missing = {"Sd_max": [], "Sd_min": [], "Pg_max": []}

for r in range(2, ws_load.max_row + 1):
    name = ws_load.cell(r, LOAD_DFES_COL).value
    if not name:          # skip empty rows
        continue
    name = str(name).strip()

    # Sd_max
    if name in max_2024:
        ws_load.cell(r, LOAD_Sd_max).value = float(max_2024[name])
    else:
        missing["Sd_max"].append(name)

    # Sd_min
    if name in min_2024:
        ws_load.cell(r, LOAD_Sd_min).value = float(min_2024[name])
    else:
        missing["Sd_min"].append(name)

# ---------------- 4. Populate the ‘gen’ sheet -------------------------
for r in range(2, ws_gen.max_row + 1):
    name = ws_gen.cell(r, GEN_DFES_COL).value
    if not name:
        continue
    name = str(name).strip()

    if name in gen_2024:
        ws_gen.cell(r, GEN_Pg_max).value = float(gen_2024[name])
    else:
        missing["Pg_max"].append(name)

# ---------------- 5. Save & report ------------------------------------
wb.save(TARGET_WB)
print(f"Updated “{TARGET_WB.name}” successfully.")

for k, v in missing.items():
    if v:
        uniq = sorted(set(v))
        print(f"⚠️  {k}: no DFES 2024 value for {len(uniq)} name(s): {', '.join(uniq[:10])}{' …' if len(uniq)>10 else ''}")
